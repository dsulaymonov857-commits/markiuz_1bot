from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook


HEADER_VALUES = {
    "code",
    "codes",
    "kod",
    "kodlar",
    "код",
    "коды",
    "datamatrix",
    "data matrix",
}


def _normalize_code(value: str) -> str:
    return (
        value.strip(" \t\r\n")
        .replace("_x001D_", "\x1d")
        .replace("_x001d_", "\x1d")
        .replace("\\u001D", "\x1d")
        .replace("\\u001d", "\x1d")
        .replace("<GS>", "\x1d")
        .replace("<gs>", "\x1d")
    )


def _unique_codes(values: list[str]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        # Do not use strip(): ASCII 29 (GS) is required inside marking codes
        # and Python considers it whitespace.
        code = _normalize_code(value)
        if not code or code.lower() in HEADER_VALUES or code in seen:
            continue
        seen.add(code)
        result.append(code)
    return result


def read_codes(file_name: str, content: bytes) -> list[str]:
    suffix = Path(file_name).suffix.lower()
    if suffix == ".csv":
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("cp1251")
        # Read physical lines without parsing delimiters. Commas and semicolons
        # can be part of the cryptographic signature and must remain untouched.
        rows = []
        # splitlines() also splits on ASCII 29 (GS), which is part of a valid
        # marking code. Split only on real CSV line endings.
        for line in text.replace("\r\n", "\n").replace("\r", "\n").split("\n"):
            value = line.strip(" \t\r\n")
            if len(value) >= 2 and value.startswith('"') and value.endswith('"'):
                value = value[1:-1].replace('""', '"')
            rows.append(value)
        return _unique_codes(rows)

    if suffix == ".xlsx":
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        values = [
            "".join(str(cell) for cell in row if cell is not None)
            for sheet in workbook.worksheets
            for row in sheet.iter_rows(values_only=True)
            if any(cell is not None for cell in row)
        ]
        workbook.close()
        return _unique_codes(values)

    raise ValueError("Faqat .xlsx yoki .csv fayl yuboring.")


def select_full_marking_codes(codes: list[str], product_type: str | None = None) -> list[str]:
    if product_type == "Mineral o'g'itlar":
        return [
            code
            for code in codes
            if code.startswith("01")
            and code[2:16].isdigit()
            and "21" in code[16:20]
            and "\x1d93" in code
        ]
    return [
        code
        for code in codes
        if code.startswith("01")
        and code[2:16].isdigit()
        and "\x1d91" in code
        and "\x1d92" in code
    ]
